from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.responses import FileResponse
from db.mongodb import get_db
from models.mentor import Mentor
import secrets
import bcrypt
from pymongo import ReturnDocument
import smtplib
from email.message import EmailMessage
import openpyxl
import io
import os
from typing import List, Dict, Any

router = APIRouter()
security = HTTPBasic()

# Portal Owner credentials (same as HOD)
PORTAL_USERNAME = "portaluser@lenovo.com"
# Hashed password for "secret123"
PORTAL_HASHED_PASSWORD = bcrypt.hashpw(b"secret123", bcrypt.gensalt())

def authenticate(credentials: HTTPBasicCredentials = Depends(security)):
    correct_username = secrets.compare_digest(credentials.username, PORTAL_USERNAME)
    correct_password = bcrypt.checkpw(credentials.password.encode(), PORTAL_HASHED_PASSWORD)

    if not (correct_username and correct_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unauthorized: Invalid Portal Owner credentials.",
            headers={"WWW-Authenticate": "Basic"},
        )

def get_next_mentor_id(mentor_collection):
    # Ensure the sequence tracker exists and starts at 2000
    mentor_collection.update_one(
        {"_id": "mentor_sequence_tracker"},
        {"$setOnInsert": {"mentor_id": 2000}},
        upsert=True
    )

    # Now increment and return the next value
    tracker = mentor_collection.find_one_and_update(
        {"_id": "mentor_sequence_tracker"},
        {"$inc": {"mentor_id": 1}},
        return_document=ReturnDocument.AFTER
    )
    return tracker["mentor_id"]

from fastapi import Request
from fastapi.templating import Jinja2Templates
templates = Jinja2Templates(directory="src/templates")

@router.post("/api/create-mentor")
async def create_mentor(request: Request, mentor: Mentor, _: HTTPBasicCredentials = Depends(authenticate)):
    mentor_collection = get_db()
    # Validate contact number length (user-friendly error)
    if not mentor.contact_number.isdigit() or len(mentor.contact_number) != 10:
        return templates.TemplateResponse("create_mentor.html", {"request": request, "error": "Contact number must be exactly 10 digits."})
    try:
        # Check for existing mentor with same email
        existing = mentor_collection.find_one({
            "email": mentor.email,
            "user_type": "mentor"
        })
        if existing:
            raise HTTPException(
                status_code=400,
                detail="Mentor with this email is already registered."
            )

        # Get next sequential mentor ID
        mentor_id = get_next_mentor_id(mentor_collection)

        mentor_data = mentor.dict()
        mentor_data["user_type"] = "mentor"
        mentor_data["mentor_id"] = mentor_id

        # Send email to mentor after successful registration
        sender_email = "bharathisriram2001@gmail.com"
        receiver_email = mentor.email
        subject = "Welcome to Mentor Portal"
        body = f"Dear Mentor,\n\nYour account has been created. Please set your password using the following link: http://127.0.0.1:8000/setup-password?email={mentor.email}\n\nYour username: {mentor.email.split('@')[0]}\n\nRegards,\nPortal Team"
        msg = EmailMessage()
        msg.set_content(body)
        msg['Subject'] = subject
        msg['From'] = sender_email
        msg['To'] = receiver_email
        mail_sent = False
        try:
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(sender_email, 'pyzs anhf fgum fkch')  # Use app password
                smtp.send_message(msg)
            mail_sent = True
        except Exception as e:
            print(f"Error sending email: {e}")
            mail_sent = False

        mentor_data["mail_sent"] = mail_sent
        result = mentor_collection.insert_one(mentor_data)

        return {
            "id": str(result.inserted_id),
            "mentor_id": mentor_id,
            "message": "Data inserted and email sent",
            "mail_sent": mail_sent
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/mentors")
async def get_mentors(_: HTTPBasicCredentials = Depends(authenticate)):
    mentor_collection = get_db()
    try:
        mentors = list(mentor_collection.find({"user_type": "mentor"}))
        for mentor in mentors:
            mentor["_id"] = str(mentor["_id"])
        return mentors
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# POC Management Endpoints

@router.post("/api/poc")
async def create_poc(poc_data: Dict[str, Any]):
    """Create a new POC entry"""
    mentor_collection = get_db()
    try:
        # Add POC to database
        poc_data["_id"] = f"poc_{len(list(mentor_collection.find({'type': 'poc'}))) + 1}"
        poc_data["type"] = "poc"
        result = mentor_collection.insert_one(poc_data)
        return {"message": "POC created successfully", "id": str(result.inserted_id)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/pocs")
async def get_pocs():
    """Get all POC entries"""
    mentor_collection = get_db()
    try:
        pocs = list(mentor_collection.find({"type": "poc"}))
        for poc in pocs:
            poc["_id"] = str(poc["_id"])
        return {"pocs": pocs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/api/poc/{poc_id}")
async def update_poc(poc_id: str, poc_data: Dict[str, Any]):
    """Update a POC entry"""
    mentor_collection = get_db()
    try:
        result = mentor_collection.update_one(
            {"_id": poc_id, "type": "poc"},
            {"$set": poc_data}
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="POC not found")
        return {"message": "POC updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/api/poc/{poc_id}")
async def delete_poc(poc_id: str):
    """Delete a POC entry"""
    mentor_collection = get_db()
    try:
        result = mentor_collection.delete_one({"_id": poc_id, "type": "poc"})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="POC not found")
        return {"message": "POC deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/poc/import-excel")
async def import_poc_excel(file: UploadFile = File(...)):
    """Import POCs from Excel file"""
    mentor_collection = get_db()
    try:
        # Read the uploaded Excel file
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents))
        worksheet = workbook.active
        
        # Expected columns: POC Title, Description, Required Skills, Difficulty Level
        expected_headers = ['POC Title', 'Description', 'Required Skills', 'Difficulty Level']
        
        # Get headers from first row
        headers = [cell.value for cell in worksheet[1]]
        
        # Validate headers
        if headers != expected_headers:
            raise HTTPException(
                status_code=400, 
                detail=f"Invalid Excel format. Expected headers: {expected_headers}, Got: {headers}"
            )
        
        imported_count = 0
        poc_data_list = []
        
        # Group rows by POC (rows with same POC Title and Description)
        poc_groups = {}

        # Process data rows (starting from row 2)
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):  # Skip completely empty rows
                continue

            poc_title, description, required_skill, difficulty_level = row

            # Validate difficulty level
            if not difficulty_level or str(difficulty_level).strip().upper() not in ['A', 'I', 'B']:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid or missing Difficulty Level at row {row_num}: '{difficulty_level}'. Must be one of A, I, B."
                )

            # Create a key for grouping (POC Title + Description)
            if poc_title and description:
                # This is a new POC entry
                poc_key = f"{str(poc_title).strip()}|{str(description).strip()}"

                difficulty_str = str(difficulty_level).strip().upper()

                # Initialize or update POC group
                if poc_key not in poc_groups:
                    poc_groups[poc_key] = {
                        "title": str(poc_title).strip(),
                        "description": str(description).strip(),
                        "required_skills": [],
                        "difficulties": [],
                        "students": []  # Keep students array but leave empty from import
                    }

                # Add skill and its corresponding difficulty (always, even if repeated)
                if required_skill:
                    skill = str(required_skill).strip()
                    poc_groups[poc_key]["required_skills"].append(skill)
                    poc_groups[poc_key]["difficulties"].append(difficulty_str)

            elif required_skill or difficulty_level:
                # This is an additional skill/difficulty for the previous POC
                # Find the last POC group to add this to
                if poc_groups:
                    last_poc_key = list(poc_groups.keys())[-1]

                    skill = str(required_skill).strip() if required_skill else ""
                    difficulty_str = str(difficulty_level).strip().upper()
                    poc_groups[last_poc_key]["required_skills"].append(skill)
                    poc_groups[last_poc_key]["difficulties"].append(difficulty_str)
        
        # Convert grouped POCs to database format
        skipped_count = 0
        for poc_data in poc_groups.values():
            # Skip empty POCs
            if not poc_data["title"] or not poc_data["description"]:
                continue

            # Check for duplicate in DB (same title and description)
            existing = mentor_collection.find_one({
                "type": "poc",
                "title": poc_data["title"],
                "description": poc_data["description"]
            })
            if existing:
                skipped_count += 1
                continue

            # Store required_skills and difficulties as parallel lists (for export)
            poc_entry = {
                "_id": f"poc_{len(poc_data_list) + len(list(mentor_collection.find({'type': 'poc'}))) + 1}",
                "type": "poc",
                "title": poc_data["title"],
                "description": poc_data["description"],
                "required_skills": poc_data["required_skills"],
                "difficulties": poc_data["difficulties"],
                "matched_students": []
            }

            poc_data_list.append(poc_entry)
            imported_count += 1
        
        # Insert all POCs into database
        if poc_data_list:
            mentor_collection.insert_many(poc_data_list)
        
        return {
            "message": f"Successfully imported {imported_count} POC entries. Skipped {skipped_count} duplicate(s).",
            "imported_count": imported_count,
            "skipped_duplicates": skipped_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

@router.get("/api/poc/export-excel")
async def export_poc_excel():
    """Export POCs to Excel file"""
    mentor_collection = get_db()
    try:
        # Get all POCs from database
        pocs = list(mentor_collection.find({"type": "poc"}))
        
        # Create new workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "POC Ideas"
        
        # Headers (excluding Matched Student for simplified Excel format)
        headers = ['POC Title', 'Description', 'Required Skills', 'Difficulty Level']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            # Style the header
            from openpyxl.styles import Font, PatternFill, Alignment
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Difficulty level reverse mapping (keep B, I, A format)
        difficulty_reverse_mapping = {
            'Beginner': 'B',
            'Intermediate': 'I',
            'Advanced': 'A',
            'B': 'B',
            'I': 'I',
            'A': 'A'
        }
        
        # Data rows (only export 4 columns)
        for row, poc in enumerate(pocs, 2):
            worksheet.cell(row=row, column=1, value=poc.get('title', ''))
            worksheet.cell(row=row, column=2, value=poc.get('description', ''))
            # Export required_skills as comma-separated
            required_skills = poc.get('required_skills')
            if isinstance(required_skills, list):
                worksheet.cell(row=row, column=3, value=", ".join(required_skills))
            else:
                worksheet.cell(row=row, column=3, value="")

            # Export difficulties as comma-separated, matching skills order
            difficulties = poc.get('difficulties')
            if isinstance(difficulties, list):
                worksheet.cell(row=row, column=4, value=", ".join(difficulties))
            else:
                worksheet.cell(row=row, column=4, value=poc.get('difficulty', 'B'))
        
        # Set column widths (4 columns only)
        worksheet.column_dimensions['A'].width = 15  # POC Title
        worksheet.column_dimensions['B'].width = 35  # Description
        worksheet.column_dimensions['C'].width = 20  # Required Skills
        worksheet.column_dimensions['D'].width = 18  # Difficulty Level
        
        # Add data validation for Difficulty Level column
        from openpyxl.worksheet.datavalidation import DataValidation
        difficulty_validation = DataValidation(
            type='list',
            formula1='"B,I,A"',
            allow_blank=False
        )
        difficulty_validation.error = 'Please select B (Beginner), I (Intermediate), or A (Advanced)'
        difficulty_validation.errorTitle = 'Invalid Difficulty Level'
        difficulty_validation.prompt = 'Enter only: B (Beginner), I (Intermediate), or A (Advanced)'
        difficulty_validation.promptTitle = 'Difficulty Level'
        
        worksheet.add_data_validation(difficulty_validation)
        difficulty_validation.add('D2:D1000')
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
        
        # Save to temporary file
        temp_file = "temp_poc_export.xlsx"
        workbook.save(temp_file)
        
        return FileResponse(
            path=temp_file,
            filename="poc_ideas_export.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting Excel file: {str(e)}")

def update_poc_matched_students(student_regno: str, skill_grades: dict, poc_collection):
    """
    For each POC, check if the student's skill grades match all required skills and difficulties.
    Extra skills in the candidate are allowed, but all required skills must be present with matching grades.
    If matched, add the student_regno to the POC's students list in the DB.
    """
    pocs = list(poc_collection.find({"type": "poc"}))
    print(f"Checking {len(pocs)} POCs for student {student_regno}...")
    for poc in pocs:
        required_skills = poc.get("required_skills", [])
        required_difficulties = poc.get("difficulties", [])
        print(f"POC '{poc.get('title', poc.get('_id'))}': Required skills={required_skills}, Required difficulties={required_difficulties}")
        if not required_skills or not required_difficulties:
            print("  Skipped: Missing required skills or difficulties.")
            continue
        match = True
        for skill, diff in zip(required_skills, required_difficulties):
            candidate_grade = skill_grades.get(skill.lower()) or skill_grades.get(skill)
            print(f"    Skill '{skill}': Candidate grade={candidate_grade}, Required={diff}")
            # Define skill levels order
            skill_levels = {'B': 1, 'I': 2, 'A': 3}
            if not candidate_grade or skill_levels.get(candidate_grade, 0) < skill_levels.get(diff, 0):
                match = False
                print("    -> Not matched.")
                break
        if match:
            students = poc.get("matched_students", [])
            if student_regno not in students:
                students.append([student_regno])
                poc_collection.update_one(
                    {"_id": poc["_id"]},
                    {"$set": {"matched_students": students}}
                )
                print(f"  -> Matched! Added student {student_regno}. Students now: {students}")
            else:
                print(f"  -> Already matched. Students: {students}")
        else:
            print(f"  -> Not matched for this POC.")
