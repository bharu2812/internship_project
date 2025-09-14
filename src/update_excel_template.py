import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl.comments
import os

def create_poc_excel_template():
    """
    Creates the complete POC Excel template with validation and formatting
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'POC Ideas'

    # Headers in row 1
    headers = ['POC Title', 'Description', 'Required Skills', 'Difficulty Level']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        # Style the header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # No sample data - keep template clean for user input

    # Add data validation for Difficulty Level column (column D) with DROPDOWN
    difficulty_validation = DataValidation(
        type='list',
        formula1='"B,I,A"',
        allow_blank=False,
        showDropDown=True,  # Enable dropdown arrow
        showErrorMessage=True,
        errorTitle='Invalid Difficulty Level',
        error='Please select B (Beginner), I (Intermediate), or A (Advanced)',
        showInputMessage=True,
        promptTitle='Difficulty Level',
        prompt='Select from dropdown: B (Beginner), I (Intermediate), or A (Advanced)'
    )

    # Apply validation to column D (starting from row 2 to row 1000)
    ws.add_data_validation(difficulty_validation)
    difficulty_validation.add('D2:D1000')

    # Add Skills guidance with hover instructions
    skills_validation = DataValidation(
        type="textLength",
        operator="greaterThan", 
        formula1="0",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
        showInputMessage=True,
        promptTitle="Required Skills",
        prompt="Enter only ONE skill here. For additional skills, enter in cells below (C3, C4, C5, etc.)"
    )

    # Apply skills guidance to Required Skills column
    skills_validation.add("C2:C1000")
    ws.add_data_validation(skills_validation)

    # Add hover comments to column headers
    # Skills column comment (C1)
    skills_comment = openpyxl.comments.Comment(
        "📋 SKILLS INSTRUCTIONS:\n\n"
        "• Enter ONLY ONE skill per cell\n"
        "• For multiple skills:\n"
        "  - First skill in C2\n"
        "  - Second skill in C3\n"
        "  - Third skill in C4\n"
        "  - And so on...\n\n"
        "Examples:\n"
        "• Python\n"
        "• JavaScript\n"
        "• Machine Learning\n"
        "• Database Design",
        "System"
    )
    ws.cell(row=1, column=3).comment = skills_comment

    # Difficulty Level column comment (D1)
    difficulty_comment = openpyxl.comments.Comment(
        "🎯 DIFFICULTY LEVEL:\n\n"
        "Use the dropdown to select:\n"
        "• B = Beginner\n"
        "• I = Intermediate\n"
        "• A = Advanced\n\n"
        "Click the dropdown arrow ▼\n"
        "to choose your option",
        "System"
    )
    ws.cell(row=1, column=4).comment = difficulty_comment

    # Set column widths for better readability (matching your image)
    ws.column_dimensions['A'].width = 15  # POC Title
    ws.column_dimensions['B'].width = 35  # Description  
    ws.column_dimensions['C'].width = 20  # Required Skills
    ws.column_dimensions['D'].width = 18  # Difficulty Level

    # Style the Difficulty Level column with light blue background
    for row in range(2, 21):  # Style empty cells for visual guidance
        cell = ws.cell(row=row, column=4)
        cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Style the Required Skills column with light green background
    for row in range(2, 21):  # Style empty cells for visual guidance
        cell = ws.cell(row=row, column=3)
        cell.fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Create static directory if it doesn't exist
    static_dir = os.path.join(os.path.dirname(__file__), 'static')
    os.makedirs(static_dir, exist_ok=True)

    # Save the file
    file_path = os.path.join(static_dir, 'poc_idea_template.xlsx')
    wb.save(file_path)
    
    return file_path

if __name__ == "__main__":
    # Create the Excel template
    template_path = create_poc_excel_template()
    print(f'✅ POC Excel template created successfully!')
    print(f'📁 Location: {template_path}')
    print('📋 Features:')
    print('   • Headers: POC Title, Description, Required Skills, Difficulty Level')
    print('   • Clean template - no sample data')
    print('   • DROPDOWN validation for Difficulty Level (B, I, A)')
    print('   • HOVER instructions for Skills section')
    print('   • Professional styling and formatting')
    print('   • Skills guidance: One skill per cell, use C3, C4 for additional skills')
    print('   • Ready for import/export functionality')
